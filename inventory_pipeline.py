"""Business logic for the self-running FG Inventory bot.

The calculations deliberately mirror the validated local FG summary builder.
This module keeps the logic callable and testable instead of executing it at
import time, which is essential for a cloud job with explicit validation gates.
"""

from __future__ import annotations

import math
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from secondary_sales import SECONDARY_OUTPUT_COLUMNS, secondary_payload


FG_SOURCE_COLUMNS = [
    "Category",
    "Brand",
    "Depot Code",
    "Depot Name",
    "SkuCode",
    "Product Name",
    "MRP",
    "Cost Price",
    "Inventory Value on CP",
    "Stock on Hand",
    "Damaged Stock",
    "Stock In Transfer",
    "Open Purchase",
    "Last 30 days Sales",
    "Day Of Inventory",
    "Last 7 days Sales",
    "Day Of Inventory2",
]

OUTPUT_COLUMNS = [
    *FG_SOURCE_COLUMNS,
    "Location Name",
    "Location type",
    "check 1",
    "Overall DRR",
    "Overall DOI",
    "Inventory Check",
    "DRR without DS",
    "DOI without DS",
    "Location DRR",
    "Location DOI",
    "Location DOI (SOH+SIT)",
    "SKU Loc SOH",
    "SKU Loc SIT",
    "Mumbai DRR",
    "Mumbai DOI",
    "3PL DRR",
    "3PL DOI",
    "Total DRR",
    "Total SOH",
    "Total DOI",
]

NUMERIC_SOURCE_COLUMNS = [
    "MRP",
    "Cost Price",
    "Inventory Value on CP",
    "Stock on Hand",
    "Damaged Stock",
    "Stock In Transfer",
    "Open Purchase",
    "Last 30 days Sales",
    "Last 7 days Sales",
]

ROUND_COLUMNS = [
    "DRR without DS",
    "DOI without DS",
    "Overall DRR",
    "Overall DOI",
    "Location DRR",
    "Location DOI",
    "Location DOI (SOH+SIT)",
    "Mumbai DRR",
    "Mumbai DOI",
    "3PL DRR",
    "3PL DOI",
    "Total DRR",
    "Total DOI",
]

BRAND_MAP = {
    "manmatters": "ManMatters",
    "man matters": "ManMatters",
    "bebodywise": "BeBodywise",
    "be bodywise": "BeBodywise",
    "littlejoys": "LittleJoys",
    "little joys": "LittleJoys",
    "root labs": "Root Labs",
    "root labs usa": "Root Labs",
    "own": "OWN",
    "stay steady": "Stay Steady",
}


def _clean_text(series: pd.Series) -> pd.Series:
    return series.fillna("").astype(str).str.strip()


def _yes(series: pd.Series) -> pd.Series:
    return _clean_text(series).str.casefold().eq("yes")


def _number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    try:
        numeric = float(value)
        return numeric if math.isfinite(numeric) else 0.0
    except (TypeError, ValueError):
        return 0.0


def normalize_brand(value: Any) -> str:
    if value is None or pd.isna(value):
        return "Unknown"
    text = str(value or "").strip()
    if not text:
        return "Unknown"
    return BRAND_MAP.get(text.casefold(), text)


def load_location_master(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        raw = pd.read_csv(path, dtype=str).fillna("")
    else:
        raw = pd.read_excel(path, dtype=str).fillna("")
    aliases = {
        "New Tagiing for Sale": "Location Name",
        "3PL": "Location type",
        "For Sale Check1": "check 1",
        "To be considered?": "Inventory Check",
    }
    master = raw.rename(columns=aliases).copy()
    override_path = path.with_name("facility_overrides.csv")
    if override_path.exists():
        overrides = pd.read_csv(override_path, dtype=str).fillna("")
        overrides = overrides.rename(columns=aliases)
        master = pd.concat(
            [
                master.loc[~master["Depot Name"].isin(overrides["Depot Name"])],
                overrides,
            ],
            ignore_index=True,
        )
    required = {
        "Depot Name",
        "Location Name",
        "Location type",
        "check 1",
        "Inventory Check",
    }
    missing = required.difference(master.columns)
    if missing:
        raise ValueError(f"Location master is missing columns: {sorted(missing)}")
    master = master[list(required)].copy()
    for column in required:
        master[column] = _clean_text(master[column])
    duplicates = master.loc[
        master["Depot Name"].duplicated(keep=False), "Depot Name"
    ].drop_duplicates()
    if not duplicates.empty:
        raise ValueError(
            "Location master has duplicate Depot Name values: "
            + ", ".join(duplicates.head(20))
        )
    return master


def _max_velocity(grouped: Any) -> pd.Series:
    return pd.concat(
        [
            grouped["Last 30 days Sales"].sum() / 30,
            grouped["Last 7 days Sales"].sum() / 7,
        ],
        axis=1,
    ).max(axis=1)


def build_inventory_frame(
    source_csv: Path,
    master_path: Path,
    *,
    min_rows: int = 18_000,
    min_skus: int = 100,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    source = pd.read_csv(source_csv, low_memory=False)
    if source.shape[1] < len(FG_SOURCE_COLUMNS):
        raise ValueError(
            f"FG source has {source.shape[1]} columns; expected at least "
            f"{len(FG_SOURCE_COLUMNS)}."
        )
    if str(source.columns[0]).strip() != "Category":
        raise ValueError(
            f"FG source first column is {source.columns[0]!r}; expected 'Category'."
        )
    source = source.iloc[:, : len(FG_SOURCE_COLUMNS)].copy()
    source.columns = FG_SOURCE_COLUMNS
    raw_rows = len(source)
    source["Depot Name"] = _clean_text(source["Depot Name"])
    source["SkuCode"] = _clean_text(source["SkuCode"])
    for column in NUMERIC_SOURCE_COLUMNS:
        source[column] = pd.to_numeric(source[column], errors="coerce").fillna(0)

    nonzero = (
        source["Stock on Hand"]
        + source["Stock In Transfer"]
        + source["Last 30 days Sales"]
        + source["Last 7 days Sales"]
    ).ne(0)
    source = source.loc[nonzero].reset_index(drop=True)
    master = load_location_master(master_path)
    frame = source.merge(master, on="Depot Name", how="left", validate="many_to_one")
    for column in [
        "Location Name",
        "Location type",
        "check 1",
        "Inventory Check",
    ]:
        frame[column] = _clean_text(frame[column])

    unmapped_rows = frame["Location Name"].eq("")
    if unmapped_rows.any():
        detail = (
            frame.loc[unmapped_rows]
            .groupby("Depot Name", dropna=False)
            .agg(lines=("SkuCode", "size"), skus=("SkuCode", "nunique"))
            .reset_index()
            .sort_values(["lines", "Depot Name"], ascending=[False, True])
        )
        evidence = "; ".join(
            f"{row['Depot Name']} ({int(row['lines'])} lines, {int(row['skus'])} SKUs)"
            for row in detail.head(20).to_dict("records")
        )
        raise ValueError(
            "Current FG source contains unmapped facilities. Publication blocked: "
            + evidence
        )

    check_sale = _yes(frame["check 1"])
    inventory_check = _yes(frame["Inventory Check"])
    location_fold = frame["Location Name"].str.casefold()
    not_dark_store = location_fold.ne("dark store")
    not_rtv = ~location_fold.str.contains("rtv", na=False)
    is_mumbai = location_fold.str.contains("mumbai", na=False)
    is_plain_mumbai = location_fold.eq("mumbai")
    is_3pl = frame["Location type"].str.casefold().eq("3pl")
    inventory_no_ds = inventory_check & not_dark_store & not_rtv
    inventory_overall = inventory_check & not_rtv

    # Without-Dark-Store network metrics.
    sale_no_ds = frame.loc[check_sale & not_dark_store].groupby("SkuCode")
    inv_no_ds = frame.loc[inventory_no_ds].groupby("SkuCode")
    drr_no_ds = _max_velocity(sale_no_ds).rename("DRR without DS")
    stock_no_ds = (
        inv_no_ds["Stock on Hand"].sum()
        + inv_no_ds["Stock In Transfer"].sum()
    ).rename("_stock_no_ds")
    frame = frame.join(drr_no_ds, on="SkuCode").join(stock_no_ds, on="SkuCode")
    frame[["DRR without DS", "_stock_no_ds"]] = frame[
        ["DRR without DS", "_stock_no_ds"]
    ].fillna(0)
    frame["DOI without DS"] = np.where(
        frame["DRR without DS"] > 0,
        frame["_stock_no_ds"] / frame["DRR without DS"],
        0,
    )
    frame.drop(columns="_stock_no_ds", inplace=True)

    # Network-wide metrics: demand from all rows; stock from eligible non-RTV rows.
    all_sale = frame.groupby("SkuCode")
    all_inventory = frame.loc[inventory_overall].groupby("SkuCode")
    overall_drr = _max_velocity(all_sale).rename("Overall DRR")
    overall_stock = (
        all_inventory["Stock on Hand"].sum()
        + all_inventory["Stock In Transfer"].sum()
    ).rename("_overall_stock")
    frame = frame.join(overall_drr, on="SkuCode").join(overall_stock, on="SkuCode")
    frame[["Overall DRR", "_overall_stock"]] = frame[
        ["Overall DRR", "_overall_stock"]
    ].fillna(0)
    frame["Overall DOI"] = np.where(
        frame["Overall DRR"] > 0,
        frame["_overall_stock"] / frame["Overall DRR"],
        0,
    )
    frame.drop(columns="_overall_stock", inplace=True)

    # Location metrics require both demand and inventory eligibility.
    by_location = frame.loc[check_sale & inventory_check].groupby(
        ["SkuCode", "Location Name"]
    )
    location_drr = _max_velocity(by_location).rename("Location DRR")
    location_soh = by_location["Stock on Hand"].sum()
    location_stock = location_soh + by_location["Stock In Transfer"].sum()
    location_doi = (
        location_soh / location_drr.replace(0, np.nan)
    ).fillna(0).rename("Location DOI")
    location_doi_sit = (
        location_stock / location_drr.replace(0, np.nan)
    ).fillna(0).rename("Location DOI (SOH+SIT)")
    frame = frame.join(location_drr, on=["SkuCode", "Location Name"])
    frame = frame.join(location_doi, on=["SkuCode", "Location Name"])
    frame = frame.join(location_doi_sit, on=["SkuCode", "Location Name"])
    frame = frame.join(
        by_location["Stock on Hand"].sum().rename("SKU Loc SOH"),
        on=["SkuCode", "Location Name"],
    )
    frame = frame.join(
        by_location["Stock In Transfer"].sum().rename("SKU Loc SIT"),
        on=["SkuCode", "Location Name"],
    )
    frame[
        [
            "Location DRR",
            "Location DOI",
            "Location DOI (SOH+SIT)",
            "SKU Loc SOH",
            "SKU Loc SIT",
        ]
    ] = frame[
        [
            "Location DRR",
            "Location DOI",
            "Location DOI (SOH+SIT)",
            "SKU Loc SOH",
            "SKU Loc SIT",
        ]
    ].fillna(0)

    # Mumbai demand includes all Mumbai nodes; displayed stock uses plain Mumbai.
    mumbai_sale = frame.loc[check_sale & is_mumbai].groupby("SkuCode")
    mumbai_drr = _max_velocity(mumbai_sale).rename("Mumbai DRR")
    mumbai_stock_group = frame.loc[check_sale & is_plain_mumbai].groupby("SkuCode")
    mumbai_stock = (
        mumbai_stock_group["Stock on Hand"].sum()
        + mumbai_stock_group["Stock In Transfer"].sum()
    )
    mumbai_doi = (
        mumbai_stock / mumbai_drr.replace(0, np.nan)
    ).fillna(0).rename("Mumbai DOI")
    frame = frame.join(mumbai_drr, on="SkuCode").join(mumbai_doi, on="SkuCode")

    # Exact 3PL matches only; "Non 3PL" must never enter this group.
    third_party = frame.loc[check_sale & is_3pl].groupby("SkuCode")
    third_party_drr = _max_velocity(third_party).rename("3PL DRR")
    third_party_stock = (
        third_party["Stock on Hand"].sum()
        + third_party["Stock In Transfer"].sum()
    )
    third_party_doi = (
        third_party_stock / third_party_drr.replace(0, np.nan)
    ).fillna(0).rename("3PL DOI")
    frame = frame.join(third_party_drr, on="SkuCode").join(
        third_party_doi, on="SkuCode"
    )

    # Historical Total metrics intentionally match the without-DS calculation.
    total_sale = frame.loc[check_sale & not_dark_store].groupby("SkuCode")
    total_inventory = frame.loc[inventory_no_ds].groupby("SkuCode")
    total_drr = _max_velocity(total_sale).rename("Total DRR")
    total_stock = (
        total_inventory["Stock on Hand"].sum()
        + total_inventory["Stock In Transfer"].sum()
    ).rename("Total SOH")
    total_doi = (
        total_stock / total_drr.replace(0, np.nan)
    ).fillna(0).rename("Total DOI")
    frame = frame.join(total_drr, on="SkuCode")
    frame = frame.join(total_stock, on="SkuCode")
    frame = frame.join(total_doi, on="SkuCode")

    for column in [
        "Mumbai DRR",
        "Mumbai DOI",
        "3PL DRR",
        "3PL DOI",
        "Total DRR",
        "Total SOH",
        "Total DOI",
    ]:
        frame[column] = frame[column].fillna(0)
    for column in ROUND_COLUMNS:
        frame[column] = frame[column].round(0).astype("int64")

    frame = frame[OUTPUT_COLUMNS].copy()
    quality = {
        "source_rows": raw_rows,
        "filtered_zero_rows": raw_rows - len(frame),
        "published_rows": len(frame),
        "distinct_skus": int(frame["SkuCode"].nunique()),
        "distinct_depots": int(frame["Depot Name"].nunique()),
        "unmapped_facilities": 0,
    }
    if quality["distinct_skus"] < min_skus or quality["published_rows"] < min_rows:
        raise ValueError(
            "FG output volume is unexpectedly low: "
            f"{quality['published_rows']:,} rows / {quality['distinct_skus']:,} SKUs."
        )
    return frame, quality


def build_freshness(
    source_csv: Path,
    master_path: Path,
    report_date: date,
    *,
    min_rows: int = 20_000,
    min_skus: int = 100,
) -> tuple[dict[str, dict[str, dict[str, float]]], dict[str, Any]]:
    master = load_location_master(master_path)
    master_location = master.set_index("Depot Name")["Location Name"]
    usecols = [
        "Facility",
        "Item Type SKU Code",
        "Inventory Type",
        "Inventory Allocation",
        "Quantity",
        "Expiry",
        "Manufacturing",
    ]
    source = pd.read_csv(source_csv, usecols=usecols, low_memory=False)
    source_rows = len(source)
    source["Facility"] = _clean_text(source["Facility"])
    source["Item Type SKU Code"] = _clean_text(source["Item Type SKU Code"])
    eligible = source.loc[
        _clean_text(source["Inventory Type"]).str.casefold().eq("good_inventory")
        & _clean_text(source["Inventory Allocation"]).str.casefold().eq("true")
    ].copy()
    eligible["Mapped Location"] = eligible["Facility"].map(master_location).fillna("")
    unmapped = eligible.loc[eligible["Mapped Location"].eq(""), "Facility"]
    if not unmapped.empty:
        counts = unmapped.value_counts()
        evidence = "; ".join(
            f"{name} ({int(count)} lines)" for name, count in counts.head(20).items()
        )
        raise ValueError(
            "Current Shelfwise source contains unmapped facilities. Publication blocked: "
            + evidence
        )
    eligible["Quantity"] = pd.to_numeric(eligible["Quantity"], errors="coerce").fillna(0)
    eligible["Manufacturing Date"] = pd.to_datetime(
        eligible["Manufacturing"], errors="coerce"
    )
    eligible["Expiry Date"] = pd.to_datetime(eligible["Expiry"], errors="coerce")
    report_timestamp = pd.Timestamp(report_date)
    eligible["Actual Life Days"] = (
        eligible["Expiry Date"] - eligible["Manufacturing Date"]
    ).dt.days
    eligible["Remaining Days"] = (
        eligible["Expiry Date"] - report_timestamp
    ).dt.days
    eligible["Pending Life"] = (
        eligible["Remaining Days"] / eligible["Actual Life Days"]
    ).where(eligible["Actual Life Days"] > 0).clip(lower=0, upper=1)
    valid = eligible["Pending Life"].notna()
    eligible["_fresh80"] = eligible["Quantity"].where(
        valid & (eligible["Pending Life"] > 0.80), 0
    )
    eligible["_freshLE80"] = eligible["Quantity"].where(
        valid & (eligible["Pending Life"] <= 0.80), 0
    )
    eligible["_missing"] = eligible["Quantity"].where(~valid, 0)
    eligible["_expired"] = eligible["Quantity"].where(
        valid & (eligible["Remaining Days"] < 0), 0
    )
    eligible["_exp30"] = eligible["Quantity"].where(
        valid & eligible["Remaining Days"].between(0, 30, inclusive="both"), 0
    )
    eligible["_exp60"] = eligible["Quantity"].where(
        valid & eligible["Remaining Days"].between(31, 60, inclusive="both"), 0
    )
    eligible["_weighted"] = (
        eligible["Quantity"] * eligible["Pending Life"]
    ).where(valid, 0)
    eligible["_valid_qty"] = eligible["Quantity"].where(valid, 0)
    grouped = (
        eligible.groupby(
            ["Item Type SKU Code", "Mapped Location"],
            dropna=False,
            as_index=False,
        )
        .agg(
            fresh80=("_fresh80", "sum"),
            freshLE80=("_freshLE80", "sum"),
            missing=("_missing", "sum"),
            expired=("_expired", "sum"),
            exp30=("_exp30", "sum"),
            exp60=("_exp60", "sum"),
            weighted=("_weighted", "sum"),
            valid_qty=("_valid_qty", "sum"),
        )
    )
    output: dict[str, dict[str, dict[str, float]]] = {}
    for row in grouped.to_dict("records"):
        sku = str(row["Item Type SKU Code"])
        location = str(row["Mapped Location"])
        valid_qty = _number(row["valid_qty"])
        output.setdefault(sku, {})[location] = {
            "fresh80": round(_number(row["fresh80"])),
            "freshLE80": round(_number(row["freshLE80"])),
            "missing": round(_number(row["missing"])),
            "expired": round(_number(row["expired"])),
            "exp30": round(_number(row["exp30"])),
            "exp60": round(_number(row["exp60"])),
            "freshPct": round(_number(row["weighted"]) / valid_qty * 100, 1)
            if valid_qty
            else 0,
        }
    quality = {
        "source_rows": source_rows,
        "eligible_rows": len(eligible),
        "eligible_quantity": round(_number(eligible["Quantity"].sum())),
        "distinct_skus": int(eligible["Item Type SKU Code"].nunique()),
        "distinct_facilities": int(eligible["Facility"].nunique()),
        "invalid_shelf_date_rows": int((~valid).sum()),
        "invalid_shelf_date_quantity": round(_number(eligible["_missing"].sum())),
        "unmapped_facilities": 0,
    }
    if quality["distinct_skus"] < min_skus or quality["eligible_rows"] < min_rows:
        raise ValueError(
            "Shelfwise eligible volume is unexpectedly low: "
            f"{quality['eligible_rows']:,} rows / {quality['distinct_skus']:,} SKUs."
        )
    return output, quality


def _freshness_for_sku(
    freshness: dict[str, dict[str, dict[str, float]]],
    sku: str,
) -> dict[str, float]:
    result = {
        "fresh80": 0,
        "freshLE80": 0,
        "missing": 0,
        "expired": 0,
        "exp30": 0,
        "exp60": 0,
        "weighted": 0.0,
        "valid": 0.0,
    }
    for values in freshness.get(sku, {}).values():
        for key in ["fresh80", "freshLE80", "missing", "expired", "exp30", "exp60"]:
            result[key] += _number(values.get(key))
        valid = _number(values.get("fresh80")) + _number(values.get("freshLE80"))
        result["weighted"] += _number(values.get("freshPct")) * valid
        result["valid"] += valid
    result["freshPct"] = round(result["weighted"] / result["valid"], 1) if result["valid"] else 0
    return result


def build_payload(
    frame: pd.DataFrame,
    freshness: dict[str, dict[str, dict[str, float]]],
    *,
    report_date: date,
    source_files: dict[str, str],
    quality: dict[str, Any],
    secondary: dict[str, Any] | None = None,
    previous_secondary: dict[str, Any] | None = None,
    previous_history: dict[str, Any] | None = None,
    history_days: int = 120,
) -> dict[str, Any]:
    sku_records: list[dict[str, Any]] = []
    for sku, sku_frame in frame.groupby("SkuCode", sort=False):
        first = sku_frame.iloc[0]
        sku_fresh = _freshness_for_sku(freshness, str(sku))
        locations = []
        for location, location_frame in sku_frame.groupby("Location Name", sort=False):
            soh = _number(location_frame["Stock on Hand"].sum())
            sit = _number(location_frame["Stock In Transfer"].sum())
            if not (soh or sit):
                continue
            drr = _number(location_frame["Location DRR"].iloc[0])
            if str(location).casefold() == "mumbai" and _number(first["Mumbai DRR"]) > 0:
                drr = _number(first["Mumbai DRR"])
            if str(location).casefold() == "dark store":
                drr = max(
                    0,
                    _number(first["Overall DRR"])
                    - _number(first["DRR without DS"]),
                )
            doi = round((soh + sit) / drr) if drr > 0 else 0
            fresh = freshness.get(str(sku), {}).get(str(location), {})
            locations.append(
                {
                    "n": str(location),
                    "s": round(soh),
                    "t": round(sit),
                    "r": round(drr),
                    "d": doi,
                    "f80": round(_number(fresh.get("fresh80"))),
                    "fle80": round(_number(fresh.get("freshLE80"))),
                    "fm": round(_number(fresh.get("missing"))),
                    "e30": round(_number(fresh.get("exp30"))),
                    "exp": round(_number(fresh.get("expired"))),
                    "fp": _number(fresh.get("freshPct")),
                }
            )
        locations.sort(key=lambda item: (-item["s"], item["n"]))
        mumbai_drr = round(_number(first["Mumbai DRR"]))
        mumbai_location = next(
            (row for row in locations if row["n"].casefold() == "mumbai"),
            None,
        )
        # Match the original bot's displayed location card exactly.  The
        # workbook metric can use a narrower eligibility mask, while the card
        # shows all plain-Mumbai rows retained in the published inventory.
        mumbai_doi = (
            round((mumbai_location["s"] + mumbai_location["t"]) / mumbai_drr)
            if mumbai_location and mumbai_drr > 0
            else 0
        )
        has_secondary = all(column in sku_frame.columns for column in SECONDARY_OUTPUT_COLUMNS)
        sku_records.append(
            {
                "code": str(sku),
                "name": "" if pd.isna(first["Product Name"]) else str(first["Product Name"]).strip(),
                "brand": normalize_brand(first["Brand"]),
                "mrp": _number(first["MRP"]),
                "cp": _number(first["Cost Price"]),
                "soh": round(_number(sku_frame["Stock on Hand"].sum())),
                "sit": round(_number(sku_frame["Stock In Transfer"].sum())),
                "dmg": round(_number(sku_frame["Damaged Stock"].sum())),
                "openPO": round(_number(sku_frame["Open Purchase"].sum())),
                "val": round(_number(sku_frame["Inventory Value on CP"].sum()), 2),
                "l30": round(_number(sku_frame["Last 30 days Sales"].sum())),
                "l7": round(_number(sku_frame["Last 7 days Sales"].sum())),
                "drr": round(_number(first["Overall DRR"])),
                "doi": round(_number(first["Overall DOI"])),
                "drrNoDS": round(_number(first["DRR without DS"])),
                "doiNoDS": round(_number(first["DOI without DS"])),
                "mumDRR": mumbai_drr,
                "mumDOI": mumbai_doi,
                "pl3DRR": round(_number(first["3PL DRR"])),
                "pl3DOI": round(_number(first["3PL DOI"])),
                "totDRR": round(_number(first["Total DRR"])),
                "totSOH": round(_number(first["Total SOH"])),
                "totDOI": round(_number(first["Total DOI"])),
                "sec7": round(_number(first.get("Secondary 7 Day Sales Units", 0))),
                "secDrr7": round(_number(first.get("Secondary DRR 7 Day", 0)), 2),
                "sec30": round(_number(first.get("Secondary 30 Day Sales Units", 0))),
                "secDrr30": round(_number(first.get("Secondary DRR 30 Day", 0)), 2),
                "secDRR": round(_number(first.get("Secondary DRR", 0)), 2),
                "secMTDUnits": round(_number(first.get("Secondary MTD Units", 0)), 2),
                "secMTDValue": round(_number(first.get("Secondary MTD Sales Value", 0)), 2),
                "secLastMonthUnits": round(
                    _number(first.get("Secondary Last Month Units", 0)), 2
                ),
                "secLastMonthValue": (
                    None
                    if not has_secondary
                    or pd.isna(first.get("Secondary Last Month Sales Value"))
                    else round(
                        _number(first.get("Secondary Last Month Sales Value", 0)), 2
                    )
                ),
                "secOverallDOI": round(
                    _number(first.get("Secondary Overall DOI", 0))
                ),
                "secMumbaiDRR": round(
                    _number(first.get("Secondary Mumbai DRR", 0)), 2
                ),
                "secMumbaiDOI": round(
                    _number(first.get("Secondary Mumbai DOI", 0))
                ),
                "fresh80": round(sku_fresh["fresh80"]),
                "freshLE80": round(sku_fresh["freshLE80"]),
                "freshMissing": round(sku_fresh["missing"]),
                "expired": round(sku_fresh["expired"]),
                "exp30": round(sku_fresh["exp30"]),
                "exp60": round(sku_fresh["exp60"]),
                "freshPct": sku_fresh["freshPct"],
                "locs": locations,
            }
        )
    sku_records.sort(key=lambda item: (-item["drr"], item["code"]))
    payload: dict[str, Any] = {
        "schemaVersion": 4,
        "reportDate": report_date.strftime("%d-%m-%Y"),
        "dateKey": report_date.isoformat(),
        "sourceFile": source_files.get("fg", ""),
        "freshnessSource": source_files.get("shelfwise", ""),
        "freshnessAvailable": True,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "rowCount": len(frame),
        "skus": sku_records,
        "quality": quality,
    }
    if secondary:
        payload["secondarySales"] = secondary_payload(secondary)
    elif isinstance(previous_secondary, dict) and previous_secondary:
        payload["secondarySales"] = previous_secondary
    history = previous_history if isinstance(previous_history, dict) else {}
    history = {
        "days": dict(history.get("days", {})),
        "meta": dict(history.get("meta", {})),
    }
    _update_history(history, payload)
    keys = sorted(history["days"])[-history_days:]
    history["days"] = {key: history["days"][key] for key in keys}
    payload["history"] = history
    return payload


def _update_history(history: dict[str, Any], payload: dict[str, Any]) -> None:
    totals = [0, 0, 0.0, 0, 0, 0]
    brands: dict[str, list[float]] = {}
    skus: dict[str, list[float]] = {}
    for item in payload["skus"]:
        values = [
            item["soh"],
            item["sit"],
            item["val"],
            item["drr"],
            item["fresh80"],
            item["freshLE80"],
        ]
        totals = [a + b for a, b in zip(totals, values)]
        brand = brands.setdefault(item["brand"], [0, 0, 0.0, 0, 0, 0])
        for index, value in enumerate(values):
            brand[index] += value
        skus[item["code"]] = [
            item["soh"],
            item["sit"],
            item["drr"],
            item["doi"],
            item["fresh80"],
            item["freshLE80"],
        ]
        history["meta"][item["code"]] = [item["name"], item["brand"]]
    totals = [round(value) for value in totals]
    brands = {
        brand: [round(value) for value in values]
        for brand, values in brands.items()
    }
    history["days"][payload["dateKey"]] = {
        "t": totals,
        "b": brands,
        "s": skus,
    }


def _excel_value(value: Any) -> Any:
    if value is None or (not isinstance(value, str) and pd.isna(value)):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, np.generic):
        return value.item()
    return value


def _style_table_sheet(sheet: Any, headers: list[str], *, freeze: str = "A2") -> None:
    header_fill = PatternFill("solid", fgColor="17324D")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    sheet.freeze_panes = freeze
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 34
    wide = {
        "Product Name": 38,
        "product_name": 38,
        "Depot Name": 25,
        "SkuCode": 24,
        "child_sku": 24,
        "Location Name": 18,
        "Location DOI (SOH+SIT)": 20,
        "Inventory Value on CP": 18,
        "channel_name": 18,
        "data_through": 14,
    }
    for index, name in enumerate(headers, start=1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = wide.get(name, 16)
        folded = name.casefold()
        if "value" in folded or name in {"MRP", "Cost Price"}:
            number_format = "#,##0.00"
        elif any(token in folded for token in ["units", "drr", "doi", "stock", "soh", "sit", "sales"]):
            number_format = "#,##0.00"
        elif "date" in folded or "through" in folded:
            number_format = "yyyy-mm-dd"
        else:
            continue
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=index).number_format = number_format


def _append_dataframe(sheet: Any, data: pd.DataFrame) -> None:
    headers = [str(column) for column in data.columns]
    sheet.append(headers)
    for values in data.itertuples(index=False, name=None):
        sheet.append([_excel_value(value) for value in values])
    _style_table_sheet(sheet, headers)


def write_summary_workbook(
    frame: pd.DataFrame,
    output_path: Path,
    secondary: dict[str, Any] | None = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "FG Inventory"
    output_columns = [
        *OUTPUT_COLUMNS,
        *[column for column in SECONDARY_OUTPUT_COLUMNS if column in frame.columns],
    ]
    sheet.append(output_columns)
    for row in frame[output_columns].itertuples(index=False, name=None):
        sheet.append(row)
    header_fill = PatternFill("solid", fgColor="17324D")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    sheet.freeze_panes = "F2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 34
    widths = {
        "Product Name": 38,
        "Depot Name": 25,
        "SkuCode": 24,
        "Location Name": 18,
        "Location DOI (SOH+SIT)": 20,
        "Inventory Value on CP": 18,
    }
    money_columns = {"MRP", "Cost Price", "Inventory Value on CP"}
    integer_columns = set(output_columns).difference(
        {
            "Category",
            "Brand",
            "Depot Code",
            "Depot Name",
            "SkuCode",
            "Product Name",
            "Location Name",
            "Location type",
            "check 1",
            "Inventory Check",
        }
    )
    for index, name in enumerate(output_columns, start=1):
        letter = get_column_letter(index)
        sheet.column_dimensions[letter].width = widths.get(name, 14)
        if name in money_columns:
            number_format = "#,##0.00"
        elif name in integer_columns:
            number_format = "#,##0"
        else:
            continue
        for column_cells in sheet.iter_cols(
            min_col=index, max_col=index, min_row=2, max_row=sheet.max_row
        ):
            for cell in column_cells:
                cell.number_format = number_format

    if secondary:
        summary_sheet = workbook.create_sheet("Secondary Sales")
        totals = secondary["totals"]
        summary_sheet.append(["Secondary Sales Summary", "Value"])
        summary_rows = [
            ("Data through", totals["dataThrough"]),
            ("Latest completed day units", totals["latestUnits"]),
            ("Latest completed day sales value", totals["latestValue"]),
            ("MTD units", totals["mtdUnits"]),
            ("MTD sales value", totals["mtdValue"]),
            (f"Last month units ({totals['lastMonth']})", totals["lastMonthUnits"]),
            (
                f"Last month sales value ({totals['lastMonth']})",
                totals["lastMonthValue"]
                if totals["lastMonthValue"] is not None
                else "Unavailable: the attachment value tab covers only 32 rolling days",
            ),
            (
                "DRR rule",
                "For every SKU-channel: MAX(7-day units / 7, 30-day units / 30); SKU DRR is the sum of channel DRRs",
            ),
            (
                "Complete monthly history",
                ", ".join((secondary.get("history") or {}).get("complete_months", []))
                or "Not available",
            ),
        ]
        for label, value in summary_rows:
            summary_sheet.append([label, _excel_value(value)])
        summary_sheet.append([])
        channel_frame = secondary["channel_summary"].copy()
        channel_headers = [str(column) for column in channel_frame.columns]
        summary_sheet.append(channel_headers)
        for values in channel_frame.itertuples(index=False, name=None):
            summary_sheet.append([_excel_value(value) for value in values])
        for cell in summary_sheet[1]:
            cell.fill = PatternFill("solid", fgColor="17324D")
            cell.font = Font(bold=True, color="FFFFFF")
        channel_header_row = len(summary_rows) + 3
        for cell in summary_sheet[channel_header_row]:
            cell.fill = PatternFill("solid", fgColor="1F6F5F")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        summary_sheet.freeze_panes = f"A{channel_header_row + 1}"
        summary_sheet.auto_filter.ref = (
            f"A{channel_header_row}:{get_column_letter(len(channel_headers))}{summary_sheet.max_row}"
        )
        summary_sheet.column_dimensions["A"].width = 34
        summary_sheet.column_dimensions["B"].width = 32
        for column in range(3, len(channel_headers) + 1):
            summary_sheet.column_dimensions[get_column_letter(column)].width = 16

        inventory_once = (
            frame.groupby("SkuCode", as_index=False)
            .first()[
                [
                    "SkuCode",
                    "Secondary Overall DOI",
                ]
            ]
        )
        sku_summary = secondary["sku_summary"].merge(
            inventory_once,
            left_on="child_sku",
            right_on="SkuCode",
            how="left",
        ).drop(columns="SkuCode")
        sku_sheet = workbook.create_sheet("SKU Secondary Sales")
        _append_dataframe(sku_sheet, sku_summary)

        sku_channel_sheet = workbook.create_sheet("SKU Channel DRR")
        _append_dataframe(sku_channel_sheet, secondary["sku_channel"])

        history = secondary.get("history") or {}
        monthly_history = history.get("monthly_channel")
        if monthly_history is not None and not monthly_history.empty:
            monthly_sheet = workbook.create_sheet("Monthly Channel History")
            _append_dataframe(monthly_sheet, monthly_history)
        daily_history = history.get("daily_channel")
        if daily_history is not None and not daily_history.empty:
            daily_sheet = workbook.create_sheet("Daily Channel History")
            _append_dataframe(daily_sheet, daily_history)
    workbook.save(output_path)
